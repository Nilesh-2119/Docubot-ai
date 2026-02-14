import os
import uuid
import json
import csv
from typing import Optional
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, delete
from fastapi import UploadFile

from app.config import get_settings
from app.models.document import Document
from app.models.embedding import Embedding
from app.models.excel_row import ExcelRow
from app.utils.file_parser import extract_text
from app.utils.text_chunker import chunk_text
from app.services.embedding_service import get_embeddings

settings = get_settings()

ALLOWED_EXTENSIONS = {"pdf", "docx", "xlsx", "csv", "txt"}
MAX_FILE_SIZE = settings.MAX_FILE_SIZE_MB * 1024 * 1024  # Convert to bytes

# File types that use structured SQL storage instead of embeddings
STRUCTURED_FILE_TYPES = {"xlsx", "csv"}


def validate_file(filename: str, file_size: int) -> Optional[str]:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext not in ALLOWED_EXTENSIONS:
        return f"File type '.{ext}' not supported. Allowed: {', '.join(ALLOWED_EXTENSIONS)}"
    if file_size > MAX_FILE_SIZE:
        return f"File size exceeds {settings.MAX_FILE_SIZE_MB}MB limit"
    return None


def _try_parse_numeric(value: str):
    """Try to convert string value to int or float for JSONB storage."""
    if not value or not value.strip():
        return value
    v = value.strip()
    # Try int first
    try:
        return int(v)
    except ValueError:
        pass
    # Try float
    try:
        return float(v)
    except ValueError:
        pass
    return value


def _parse_xlsx_to_rows(file_path: str) -> list[dict]:
    """Parse XLSX file into list of {sheet_name, row_number, row_data} dicts."""
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True)
    rows = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        headers = []

        for row_num, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            row_values = [str(cell) if cell is not None else "" for cell in row]

            # First row = headers
            if row_num == 1:
                headers = [v.strip() if v.strip() else f"Column{i+1}" for i, v in enumerate(row_values)]
                continue

            # Skip empty rows
            if not any(v.strip() for v in row_values):
                continue

            # Build row_data dict with numeric conversion
            row_data = {}
            for header, value in zip(headers, row_values):
                row_data[header] = _try_parse_numeric(value)

            rows.append({
                "sheet_name": sheet_name,
                "row_number": row_num,
                "row_data": row_data,
            })

    wb.close()
    return rows


def _parse_csv_to_rows(file_path: str) -> list[dict]:
    """Parse CSV file into list of {sheet_name, row_number, row_data} dicts."""
    rows = []

    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        reader = csv.reader(f)
        headers = []

        for row_num, row in enumerate(reader, start=1):
            row_values = [cell.strip() for cell in row]

            # First row = headers
            if row_num == 1:
                headers = [v if v else f"Column{i+1}" for i, v in enumerate(row_values)]
                continue

            # Skip empty rows
            if not any(v for v in row_values):
                continue

            # Build row_data dict with numeric conversion
            row_data = {}
            for header, value in zip(headers, row_values):
                row_data[header] = _try_parse_numeric(value)

            rows.append({
                "sheet_name": "Sheet1",
                "row_number": row_num,
                "row_data": row_data,
            })

    return rows


async def upload_and_process_document(
    db: AsyncSession,
    chatbot_id: str,
    file: UploadFile,
) -> Document:
    # Read file content
    content = await file.read()
    file_size = len(content)

    # Validate
    error = validate_file(file.filename, file_size)
    if error:
        raise ValueError(error)

    ext = file.filename.rsplit(".", 1)[-1].lower()

    # Save file locally
    os.makedirs(settings.UPLOAD_DIR, exist_ok=True)
    file_id = str(uuid.uuid4())
    file_path = os.path.join(settings.UPLOAD_DIR, f"{file_id}.{ext}")

    with open(file_path, "wb") as f:
        f.write(content)

    # Create document record
    doc = Document(
        chatbot_id=chatbot_id,
        filename=file.filename,
        file_type=ext,
        file_size=file_size,
        file_path=file_path,
        status="processing",
    )
    db.add(doc)
    await db.flush()

    try:
        if ext in STRUCTURED_FILE_TYPES:
            # ═══════════════════════════════════════════════════
            # STRUCTURED SQL PATH — Excel/CSV → JSONB rows
            # No embeddings generated. Data queried via SQL.
            # ═══════════════════════════════════════════════════
            if ext == "xlsx":
                parsed_rows = _parse_xlsx_to_rows(file_path)
            else:
                parsed_rows = _parse_csv_to_rows(file_path)

            if not parsed_rows:
                doc.status = "error"
                doc.chunk_count = 0
                await db.flush()
                raise ValueError("No data could be extracted from the spreadsheet")

            # Store each row as an ExcelRow record
            for row in parsed_rows:
                db.add(ExcelRow(
                    chatbot_id=chatbot_id,
                    document_id=doc.id,
                    sheet_name=row["sheet_name"],
                    row_number=row["row_number"],
                    row_data=row["row_data"],
                ))

            doc.chunk_count = len(parsed_rows)
            doc.status = "ready"
            await db.flush()

        else:
            # ═══════════════════════════════════════════════════
            # EXISTING RAG PATH — PDF/DOCX/TXT → chunk + embed
            # Completely unchanged from original implementation.
            # ═══════════════════════════════════════════════════
            text = extract_text(file_path, ext)

            if not text.strip():
                doc.status = "error"
                doc.chunk_count = 0
                await db.flush()
                raise ValueError("No text could be extracted from the document")

            chunks = chunk_text(text)

            # Generate embeddings in batches
            batch_size = 50
            all_embeddings = []
            for i in range(0, len(chunks), batch_size):
                batch = chunks[i : i + batch_size]
                batch_embeddings = await get_embeddings(batch)
                all_embeddings.extend(batch_embeddings)

            # Store embeddings
            for i, (chunk, emb) in enumerate(zip(chunks, all_embeddings)):
                meta = json.dumps({
                    "chunk_index": i,
                    "source": file.filename,
                })
                embedding_record = Embedding(
                    chatbot_id=chatbot_id,
                    document_id=doc.id,
                    content=chunk,
                    embedding=emb,
                    metadata_json=meta,
                )
                db.add(embedding_record)

            doc.chunk_count = len(chunks)
            doc.status = "ready"
            await db.flush()

    except Exception as e:
        doc.status = "error"
        await db.flush()
        raise e

    return doc


async def get_documents(db: AsyncSession, chatbot_id: str) -> list[Document]:
    result = await db.execute(
        select(Document).where(Document.chatbot_id == chatbot_id).order_by(Document.created_at.desc())
    )
    return list(result.scalars().all())


async def delete_document(db: AsyncSession, document_id: str, chatbot_id: str) -> bool:
    result = await db.execute(
        select(Document).where(Document.id == document_id, Document.chatbot_id == chatbot_id)
    )
    doc = result.scalar_one_or_none()
    if not doc:
        return False

    # Delete file from disk
    if os.path.exists(doc.file_path):
        os.remove(doc.file_path)

    # Delete embeddings (for PDF/DOCX/TXT)
    await db.execute(
        delete(Embedding).where(Embedding.document_id == document_id)
    )

    # Delete excel rows (for XLSX/CSV)
    await db.execute(
        delete(ExcelRow).where(ExcelRow.document_id == document_id)
    )

    # Delete document
    await db.delete(doc)
    await db.flush()
    return True


async def get_embedding_count(db: AsyncSession, chatbot_id: str) -> int:
    result = await db.execute(
        select(func.count(Embedding.id)).where(Embedding.chatbot_id == chatbot_id)
    )
    return result.scalar() or 0
